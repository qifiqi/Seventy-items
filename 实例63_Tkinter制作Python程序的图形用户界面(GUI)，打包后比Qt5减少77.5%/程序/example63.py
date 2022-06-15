from os import listdir,getcwd
from xlrd import open_workbook, xldate
from datetime import datetime
from time import time, localtime ,strftime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from tkinter import Tk,Entry,Button,Listbox,X,Y,END,Scrollbar,RIGHT,BOTTOM,HORIZONTAL
from tkinter.filedialog import askdirectory

class MainGUI():
    def __init__(self):
        myWindow = Tk()
        myWindow.title("领料记录汇总")
        #设置窗口大小
        myWindow.geometry('590x400')
        myWindow.iconbitmap(getcwd()+"\\PO.ico")
        #增加文本框
        self.input_entry = Entry(myWindow, highlightcolor='red', highlightthickness=1)
        self.input_entry.place(x=10, y=10,width=480, height=30)
        self.btn_in = Button(myWindow, text='输入文件目录',command = self.select_dir1, width=10, height=1) 
        self.btn_in.place(x=500,y=10)

        self.output_entry = Entry(myWindow, highlightcolor='blue', highlightthickness=1)
        self.output_entry.place(x=10, y=50,width=480, height=30)
        self.btn_out = Button(myWindow, text='输出文件目录',command = self.select_dir2, width=10, height=1)
        self.btn_out.place(x=500,y=50)
        
        self.btn_run = Button(myWindow, text='执行汇总', width=10, height=1,command = self.Summary_data)
        self.btn_run.place(x=500,y=90)
        #增加列表框
        self.result_show = Listbox(myWindow,bg='DarkSeaGreen') #yscrollcommand = scroll_bar,
        self.result_show.place(x=10,y=130, width=570, height=260)
        self.sbY = Scrollbar(self.result_show,command=self.result_show.yview)#在列表框中增加Y轴滚动条
        self.sbY.pack(side=RIGHT,fill=Y)
        self.result_show.config(yscrollcommand = self.sbY.set)
        self.sbX = Scrollbar(self.result_show,command=self.result_show.xview,orient = HORIZONTAL)#在列表框中增加X轴滚动条
        self.sbX.pack(side=BOTTOM,fill=X)
        self.result_show.config(xscrollcommand = self.sbX.set)
        
        myWindow.mainloop()
        
    def select_dir1(self):
        self.input_entry.delete(0, END)
        self.input_entry.insert(0, askdirectory(initialdir= "D:\\"))

    def select_dir2(self):
        self.output_entry.delete(0, END)
        self.output_entry.insert(0, askdirectory(initialdir= "D:\\"))
        
    #读取xls文件中的数据
    def Get_data(self,file):
        wb = open_workbook(file) #读取工作簿
        ws = wb.sheets()[0] #选第一个工作表
        data = {}
        for row in range(7, ws.nrows-2):
            dept = ws.cell(2, 16).value #部门
            dept_id = ws.cell(3, 16).value #部门编号
            dt = ws.cell(row, 0).value #时间
            if type(dt) is float:
                date_time = xldate.xldate_as_datetime(dt, 0)
            else:
                date_time = datetime.strptime(dt,'%Y-%m-%d %H:%M:%S')
            business = ws.cell(row, 2).value #业务类型
            model = ws.cell(row, 3).value #品种
            qty = ws.cell(row, 4).value #数量
            unit_price = ws.cell(row, 6).value #单价
            price = ws.cell(row, 8).value #总价
            reward = ws.cell(row, 9).value #额外值
            discount = ws.cell(row, 11).value #调整
            balance = ws.cell(row, 13).value #剩余
            location = str(ws.cell(row, 15).value).strip() #库位
            operator = ws.cell(row, 17).value #操作员
            date = date_time.date() #日期
            time = date_time.time() #时间
            info_list=[dept,dept_id,date_time,business,model,qty,unit_price,price,reward,discount,
                       balance,location,operator,date,time]
            data.setdefault(date,[]) #以日期为键
            if info_list[3] != "备注": #不要业务类型为“备注”的数据
                data[date].append(info_list)
        #增加当日领取次数        
        for key in data.keys():
            for i in data[key]:
                i.append(len(data[key]))

        return data

    def Get_file_path(self,path):        
            files=[]
            for file in listdir(path):
                if file.endswith(".xls"): #排除文件夹内的其它干扰文件
                    files.append(path+"\\"+file)
            return files

    def Get_current_time(self):
        time_stamp = time()  # 当前时间的时间戳
        local_time = localtime(time_stamp)  #
        str_time = strftime('%Y-%m-%d %H.%M.%S', local_time)
        return str_time

    def Summary_data(self):
        thin = Side(border_style="thin", color="000000")#定义边框粗细及颜色
        title = ['部门', '部门编号', '时间', '业务类型', '品种', '数量', '单价', '金额', '额外值',
         '调整', '剩余', '库位', '操作员', '领取日期', '领取时间', '领取次数']

        wb = Workbook() 
        ws = wb.active
        ws.merge_cells("A1:P1")
        ws.cell(1,1).value = "领料明细汇总表"
        ws.cell(1,1).font = Font(name=u'黑体',bold=True,size=18)
        ws.row_dimensions[1].height  = 22.2
        ws.cell(1,1).alignment = Alignment(horizontal="center", vertical="center")
        ws.append(title)

        #插入数据
        files = self.Get_file_path(self.input_entry.get()) #get()获取文本编辑框中的输入文件目录，并获取目录下的xls文件
        for file in files:
            data = self.Get_data(file)
            for key in data.keys():
                for i in data[key]:
                    ws.append(i)
            f = f"{file} 的内容已加入总表." # 创建一个显示项
            self.result_show.insert("end", f) #将结果添加到列表框中

        #设置字号，对齐，缩小字体填充，加边框
        #Font(bold=True)可加粗字体
        for row_number in range(2, ws.max_row+1):
            for col_number in range(1,17):
                c = ws.cell(row=row_number,column=col_number)
                c.font = Font(size=9)
                c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                c.alignment = Alignment(horizontal="left", vertical="center")

        col_name= list("ABCDEFGHIJKLMNOP")
        col_width = [8, 8, 16, 8, 16, 8, 8, 9.8, 8, 8, 8, 11, 8.3, 9, 8, 8]
        for i in range(len(col_name)):
            ws.column_dimensions[col_name[i]].width = col_width[i]

        ws.column_dimensions.group('I','K',hidden=True)
        ws.column_dimensions.group('N','O',hidden=True)

        wb.save(f"{self.output_entry.get()}\\领料明细汇总表{self.Get_current_time()}.xlsx")
        f = "-"*100 #创建分割线
        self.result_show.insert("end", f) # 将分割线添加到列表框
        f = f"领料明细汇总表{self.Get_current_time()}.xlsx 已生成，请去输出文件夹查看."# 创建一个显示项
        self.result_show.insert("end", f) # 将结果添加到列表框
        f = " "*100
        self.result_show.insert("end", f) # 将以上空格添加到列表框

if __name__ == "__main__":
    MainGUI()